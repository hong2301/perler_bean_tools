import os
import json
import csv
import cv2
import numpy as np
import threading
import time
from flask import Flask, render_template, request, jsonify, send_file, Response, stream_with_context
import json as json_lib
from werkzeug.utils import secure_filename
from paddleocr import PaddleOCR
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import uuid

# 创建 Flask 应用
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = './uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB 最大文件大小

# 确保上传目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# 全局变量：OCR 引擎和颜色映射
ocr = None
color_mapping = {}
ocr_initialized = False  # OCR 初始化状态标志

# 配置参数
recognitionThreshold = 0.7
overlapValue = 10
heightValue = 400


def init_ocr():
    """初始化 OCR 引擎（只会执行一次）"""
    global ocr, ocr_initialized
    
    # 检查是否已初始化
    if ocr_initialized and ocr is not None:
        print("PaddleOCR 模型已加载，跳过重复初始化")
        return
    
    print("正在加载 PaddleOCR 模型，请稍候...")
    ocr = PaddleOCR(
        use_doc_orientation_classify=False,
        use_doc_unwarping=False,
        use_textline_orientation=False,
    )
    ocr_initialized = True
    print("PaddleOCR 模型加载完成！")


def load_color_mapping(json_path='color.json'):
    """加载颜色映射表"""
    global color_mapping
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        color_plates = data.get('color_plates', {})
        for plate_number, colors in color_plates.items():
            for color_info in colors:
                code = color_info.get('code')
                name = color_info.get('name')
                if code:
                    color_mapping[code] = (plate_number, name)
        print(f"颜色映射表加载完成，共 {len(color_mapping)} 个颜色")
    except Exception as e:
        print(f"读取颜色配置文件出错: {e}")


def split_image(image, chunk_width=0, chunk_height=1000, overlap_width=0, overlap_height=100):
    """将图片按宽度和高度分块
    
    Args:
        image: OpenCV 图像
        chunk_width: 分块宽度，0 表示使用图片完整宽度（不分宽度块）
        chunk_height: 分块高度
        overlap_width: 宽度方向重叠
        overlap_height: 高度方向重叠
    
    Returns:
        chunks: [(chunk_image, x_offset, y_offset), ...]
    """
    height, width = image.shape[:2]
    chunks = []
    
    # 如果 chunk_width 为 0 或大于等于图片宽度，则不分宽度块
    if chunk_width <= 0 or chunk_width >= width:
        chunk_width = width
        overlap_width = 0  # 不分块时不使用重叠
    
    # 计算宽度方向的分块（列）
    x_positions = []
    start_x = 0
    while start_x < width:
        end_x = min(start_x + chunk_width, width)
        x_positions.append((start_x, end_x))
        if end_x >= width:
            break
        start_x = end_x - overlap_width
    
    # 计算高度方向的分块（行）
    y_positions = []
    start_y = 0
    while start_y < height:
        end_y = min(start_y + chunk_height, height)
        y_positions.append((start_y, end_y))
        if end_y >= height:
            break
        start_y = end_y - overlap_height
    
    # 生成所有分块
    for y_start, y_end in y_positions:
        for x_start, x_end in x_positions:
            chunk = image[y_start:y_end, x_start:x_end]
            chunks.append((chunk, x_start, y_start))
    
    return chunks


def ocr_chunk(chunk, x_offset, y_offset, index):
    """对单块图片进行 OCR 识别
    
    Args:
        chunk: 图像块
        x_offset: x 方向偏移量
        y_offset: y 方向偏移量
        index: 块索引
    """
    temp_path = f"./temp_chunk{index}.jpg"
    cv2.imwrite(temp_path, chunk)

    try:
        result = ocr.predict(temp_path)
        if isinstance(result, (list, tuple)) and len(result) > 0:
            res = result[0]
            if res is not None:
                if 'det_boxes' in res:
                    for box in res['det_boxes']:
                        box[0] += x_offset  # x1
                        box[1] += y_offset  # y1
                        box[2] += x_offset  # x2
                        box[3] += y_offset  # y2
                return res
    except Exception as e:
        print(f"识别块时出错 (x_offset={x_offset}, y_offset={y_offset}): {e}")

    return None


def merge_results(results):
    """合并多个分块的 OCR 结果"""
    merged = {
        'input_path': [],
        'page_index': [],
        'det_polygons': [],
        'det_boxes': [],
        'rec_texts': [],
        'rec_scores': [],
        'rec_boxes': []
    }

    for res in results:
        if res is None:
            continue
        for key in merged.keys():
            if key in res:
                if isinstance(res[key], list):
                    merged[key].extend(res[key])
                else:
                    merged[key].append(res[key])

    return merged


def cleanup_temp_chunks():
    """清理临时分块图片文件"""
    try:
        for filename in os.listdir('.'):
            if filename.startswith('temp_chunk') and filename.endswith('.jpg'):
                try:
                    os.remove(filename)
                    print(f"已删除临时文件: {filename}")
                except Exception as e:
                    print(f"删除临时文件失败 {filename}: {e}")
    except Exception as e:
        print(f"清理临时文件时出错: {e}")


def ocr_with_chunks(image_path, chunk_width=0, chunk_height=1000, overlap_width=0, overlap_height=100, progress_callback=None):
    """分块 OCR 主函数
    
    Args:
        image_path: 图片路径
        chunk_width: 分块宽度，0 表示使用完整宽度
        chunk_height: 分块高度
        overlap_width: 宽度方向重叠
        overlap_height: 高度方向重叠
        progress_callback: 进度回调函数，参数为 (percent, message)
    """
    image = cv2.imread(image_path)

    if image is None:
        return None, "无法读取图片"

    height, width = image.shape[:2]

    # 如果不需要分块，直接识别
    if (chunk_width <= 0 or chunk_width >= width) and height <= chunk_height:
        if progress_callback:
            progress_callback(50, '正在进行OCR识别...')
        result = ocr.predict(image_path)
        if progress_callback:
            progress_callback(100, '识别完成！')
        if isinstance(result, (list, tuple)) and len(result) > 0:
            return result[0], None
        return None, "OCR 识别失败"

    chunks = split_image(image, chunk_width, chunk_height, overlap_width, overlap_height)
    results = []
    total_chunks = len(chunks)

    for i, (chunk, x_offset, y_offset) in enumerate(chunks):
        if progress_callback:
            percent = int((i / total_chunks) * 80)  # 识别阶段占80%
            progress_callback(percent, f'正在识别第 {i+1}/{total_chunks} 块...')
        
        res = ocr_chunk(chunk, x_offset, y_offset, i + 1)
        if res:
            results.append(res)

    if progress_callback:
        progress_callback(90, '正在合并识别结果...')
    
    merged_result = merge_results(results)
    
    # 清理临时分块文件
    cleanup_temp_chunks()
    
    if progress_callback:
        progress_callback(100, '识别完成！')
    
    return merged_result, None


def process_image(image_path, chunk_width=0, chunk_height=heightValue, overlap_width=0, overlap_height=overlapValue, progress_callback=None):
    """处理图片并返回颜色统计结果"""
    ocrResult, error = ocr_with_chunks(image_path, chunk_width, chunk_height, overlap_width, overlap_height, progress_callback)

    if error:
        return None, error

    color_count = {}

    if ocrResult:
        rec_texts = ocrResult.get('rec_texts', [])
        rec_scores = ocrResult.get('rec_scores', [])
        filtered_texts = []

        for text, score in zip(rec_texts, rec_scores):
            if score > recognitionThreshold:
                filtered_texts.append(text)

        # 处理 filtered_texts
        for item in filtered_texts:
            if not item or not item[0].isupper():
                continue

            parts = item.split()

            for part in parts:
                if not part or not part[0].isupper():
                    continue

                letter_part = ''
                number_part = ''
                for char in part:
                    if char.isalpha():
                        letter_part += char
                    elif char.isdigit():
                        number_part += char
                    else:
                        break

                if letter_part and number_part:
                    try:
                        number = int(number_part)
                        if 1 <= number <= 32:
                            color_code = f"{letter_part}{number}"
                            color_count[color_code] = color_count.get(color_code, 0) + 1
                    except ValueError:
                        continue

    return color_count, None


def generate_csv_data(color_count):
    """生成 CSV 数据"""
    data = []
    for color_code in color_count.keys():
        count = color_count[color_code]
        plate_info = color_mapping.get(color_code, ('未知', '未知'))
        plate_number = plate_info[0]
        color_name = plate_info[1]
        data.append((color_code, plate_number, count, color_name))

    # 按色盘排序
    def sort_key(item):
        plate = item[1]
        try:
            plate_int = int(plate)
        except ValueError:
            plate_int = 999
        return (plate_int, item[0])

    data.sort(key=sort_key)
    return data


def generate_excel(data, output_path):
    """生成 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "颜色统计"

    # 设置表头样式
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 写入表头
    headers = ['序号', '颜色编号', '色盘', '数量', '颜色名称']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 设置边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入数据
    for idx, row_data in enumerate(data, start=1):
        ws.cell(row=idx+1, column=1, value=idx)
        ws.cell(row=idx+1, column=2, value=row_data[0])
        ws.cell(row=idx+1, column=3, value=row_data[1])
        ws.cell(row=idx+1, column=4, value=row_data[2])
        ws.cell(row=idx+1, column=5, value=row_data[3])

    # 设置所有单元格边框和对齐
    for row in ws.iter_rows(min_row=1, max_row=len(data)+1, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 15

    wb.save(output_path)


@app.route('/')
def index():
    """首页"""
    return render_template('index.html')


# 全局变量存储处理进度
processing_progress = {}

@app.route('/upload', methods=['POST'])
def upload():
    """处理图片上传，支持流式进度"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '没有文件'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': '没有选择文件'})

    def generate():
        if file:
            filename = str(uuid.uuid4()) + '.jpg'
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            yield json_lib.dumps({'type': 'progress', 'percent': 5, 'message': '图片已保存，开始处理...'}) + '\n'

            # 获取分块参数
            chunk_width = request.form.get('chunk_width', type=int, default=0)
            chunk_height = request.form.get('chunk_height', type=int, default=heightValue)
            overlap_width = request.form.get('overlap_width', type=int, default=0)
            overlap_height = request.form.get('overlap_height', type=int, default=overlapValue)
            
            yield json_lib.dumps({'type': 'progress', 'percent': 10, 'message': '准备分块识别...'}) + '\n'

            # 创建进度队列
            progress_queue = []
            
            def progress_callback(percent, message):
                # OCR识别进度映射到 10-90%
                mapped_percent = 10 + int(percent * 0.8)
                progress_queue.append({'type': 'progress', 'percent': mapped_percent, 'message': message})

            # 在线程中运行处理
            result_container = {}
            
            def process_thread():
                color_count, error = process_image(filepath, chunk_width=chunk_width, chunk_height=chunk_height, 
                                                  overlap_width=overlap_width, overlap_height=overlap_height,
                                                  progress_callback=progress_callback)
                result_container['color_count'] = color_count
                result_container['error'] = error
            
            thread = threading.Thread(target=process_thread)
            thread.start()
            
            # 等待处理完成，同时发送进度
            last_progress_count = 0
            while thread.is_alive() or len(progress_queue) > last_progress_count:
                while last_progress_count < len(progress_queue):
                    yield json_lib.dumps(progress_queue[last_progress_count]) + '\n'
                    last_progress_count += 1
                time.sleep(0.1)
            
            # 发送剩余进度
            while last_progress_count < len(progress_queue):
                yield json_lib.dumps(progress_queue[last_progress_count]) + '\n'
                last_progress_count += 1
            
            thread.join()
            
            color_count = result_container.get('color_count')
            error = result_container.get('error')

            if error:
                yield json_lib.dumps({'type': 'result', 'success': False, 'error': error}) + '\n'
                return

            yield json_lib.dumps({'type': 'progress', 'percent': 95, 'message': '正在生成结果文件...'}) + '\n'

            # 生成 CSV 数据
            csv_data = generate_csv_data(color_count)

            # 保存 CSV 文件
            csv_filename = 'openMe.csv'
            csv_path = os.path.join(app.config['UPLOAD_FOLDER'], csv_filename)

            with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(['序号', '颜色编号', '色盘', '数量', '颜色名称'])
                for idx, row in enumerate(csv_data, start=1):
                    writer.writerow([idx] + list(row))

            # 生成 Excel 文件
            xlsx_filename = 'openMe.xlsx'
            xlsx_path = os.path.join(app.config['UPLOAD_FOLDER'], xlsx_filename)
            generate_excel(csv_data, xlsx_path)

            yield json_lib.dumps({
                'type': 'result',
                'success': True,
                'data': csv_data,
                'csv_url': f'/download/{csv_filename}',
                'xlsx_url': f'/download/{xlsx_filename}'
            }) + '\n'

    return Response(stream_with_context(generate()), mimetype='application/x-ndjson')


@app.route('/query', methods=['POST'])
def query():
    """查询颜色编号对应的色盘"""
    input_text = request.json.get('colors', '').strip()

    if not input_text:
        return jsonify({'success': False, 'error': '输入为空'})

    color_codes = input_text.split()
    results = []

    for code in color_codes:
        code_upper = code.upper()
        plate = color_mapping.get(code_upper, '未找到')
        results.append({'color': code, 'plate': plate})

    return jsonify({'success': True, 'results': results})


@app.route('/download/<filename>')
def download(filename):
    """下载 CSV 文件"""
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    return send_file(filepath, as_attachment=True)


if __name__ == '__main__':
    # 启动时加载 OCR 和颜色映射
    init_ocr()
    load_color_mapping()

    # 启动 Flask 应用（使用 8080 端口避免冲突）
    # use_reloader=False 防止自动重载导致 OCR 重复加载
    app.run(debug=True, host='0.0.0.0', port=8080, use_reloader=False)
